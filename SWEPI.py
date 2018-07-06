import os
import openpyxl
from openpyxl import load_workbook
import csv
from openpyxl import Workbook
import datetime
import operator
from operator import itemgetter
from decimal import *

tags = []
sheets = []
sheet_names = []
whole_sheets = []
sheet_total_index = []
property_row_list = []
value_rows = []
headings = ["Production Date", "Product Disposition", "Acct Ref", "Typ Int", "Owner Decimal", "UOM",
            "Gross Vol", "Owner Volume", "Amount Description", "Unit Value", "Gross Amounts", "Owner Amounts",
            "Owner Net Amount"]

col_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
full_col_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']

os.chdir('/Users/porte/Desktop')
cwd = os.getcwd()
# os.unlink('/Users/mbpsmac/Desktop/try.xlsx')


# file_name = input("Please enter your file name: ")
# print("Processing file...........................")

file_name = 'swepi.csv'
#
#
file = open(file_name, 'r')
csv_f = csv.reader(file)
# clean_rows = [row for row in csv_f if "SUBTOTAL" not in row]

with open("swepioutput.csv", "w", newline='') as f:
    writer = csv.writer(f, quoting=csv.QUOTE_NONNUMERIC)
    writer.writerows(csv_f)

# wb = load_workbook(filename='new_swepi.xlsx')
# ws = wb.active

wb = Workbook()
ws = wb.active

with open('swepioutput.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)


# Captures all cells in a range, creating and returning a sheet
def cell_range(start, end):
    sheet = []
    for cell_row in ws.iter_rows(min_row=start, max_col=13, max_row=end):
        for data in cell_row:
            sheet.append(data)

    return sheet


# Captures all cell values in range
def cell_range_values(start, end):
    sheet = []
    for cell_row in ws.iter_rows(min_row=start, max_col=13, max_row=end):
        for data in cell_row:
            sheet.append(data.value)

    return sheet


def cell_range_paster(start, end, list):
    for cell_row in ws.iter_rows(min_row=start, max_col=13, max_row=end):
        for idx, cell in enumerate(cell_row):
            cell.value = list[idx].value


# Creates a list of cell values from a cell range
def cell_copier(cell_range):
    copied_cells = []
    for row in cell_range:
        for cell in row:
            copied_cells.append(cell.value)

    return copied_cells


# Deletes a range of cells
def cell_delete(cell_range):
    for row in cell_range:
        for cell in row:
            cell.value = None


# Pastes the values from a list into a range of cells
def cell_paster(cell_list, cell_range):
    counter = 0
    for row in cell_range:
        for cell in row:
            cell.value = cell_list[counter]
            counter += 1


# copies cell values from on Range, deletes the cells in that range, then pastes the values in the given range
def cell_copy_paste(copy_range, paste_range):
    counter = 0
    cell_list = cell_copier(copy_range)
    cell_delete(copy_range)
    for row in paste_range:
        for cell in row:
            cell.value = cell_list[counter]
            counter += 1


grand_total = 0

for row in ws['A']:
    row_values = []
    try:
        if "PROPERTY" in row.value:

            for col in col_list:
                try:
                    cell = ws[col + str(row.row)].value.replace("FIELD:", "").strip()
                    row_values.append(cell)
                except AttributeError:
                    cell = ""
            row_values = [x for x in row_values if x != ""]

            if len(row_values) == 5:
                row_values.pop(2)
                property_row_list.append(row_values)

            ws['A' + str(row.row)] = row_values[0]
            ws['D' + str(row.row)] = row_values[1]
            ws['G' + str(row.row)] = ""
            ws['J' + str(row.row)] = ""
            ws['B' + str(row.row)] = ""
            ws['C' + str(row.row)] = ""
            ws['E' + str(row.row)] = ""
            ws['F' + str(row.row)] = ""
            ws['H' + str(row.row)] = ""
            ws['I' + str(row.row)] = ""

        elif row.value != "":

            for col in col_list:
                try:
                    cell = ws[col + str(row.row)].value.split(" ")
                    row_values.extend(cell)
                except AttributeError:
                    cell = ""
                    row_values.extend(cell)

            row_values = [x for x in row_values if x != ""]

            if len(row_values) == 13:
                ws['A' + str(row.row)] = row_values[0]
                ws['B' + str(row.row)] = row_values[1]
                ws['C' + str(row.row)] = row_values[2]
                ws['D' + str(row.row)] = row_values[3]
                ws['E' + str(row.row)] = row_values[4]
                ws['F' + str(row.row)] = row_values[5]
                ws['G' + str(row.row)] = row_values[6]
                ws['H' + str(row.row)] = row_values[7]
                ws['I' + str(row.row)] = row_values[8]
                ws['J' + str(row.row)] = row_values[9]
                ws['K' + str(row.row)] = row_values[10]
                ws['L' + str(row.row)] = row_values[11]
                ws['M' + str(row.row)] = row_values[12]

            elif len(row_values) == 14:
                product = row_values[1] + " " + row_values[2]
                row_values[1] = product
                ws['A' + str(row.row)] = row_values[0]
                ws['B' + str(row.row)] = product
                ws['C' + str(row.row)] = row_values[3]
                ws['D' + str(row.row)] = row_values[4]
                ws['E' + str(row.row)] = row_values[5]
                ws['F' + str(row.row)] = row_values[6]
                ws['G' + str(row.row)] = row_values[7]
                ws['H' + str(row.row)] = row_values[8]
                ws['I' + str(row.row)] = row_values[9]
                ws['J' + str(row.row)] = row_values[10]
                ws['K' + str(row.row)] = row_values[11]
                ws['L' + str(row.row)] = row_values[12]
                ws['M' + str(row.row)] = row_values[13]
                row_values.pop(2)

            elif len(row_values) == 15:
                product = row_values[1] + " " + row_values[2] + " " + row_values[3]
                ws['A' + str(row.row)] = row_values[0]
                ws['B' + str(row.row)] = product
                ws['C' + str(row.row)] = row_values[4]
                ws['D' + str(row.row)] = row_values[5]
                ws['E' + str(row.row)] = row_values[6]
                ws['F' + str(row.row)] = row_values[7]
                ws['G' + str(row.row)] = row_values[8]
                ws['H' + str(row.row)] = row_values[9]
                ws['I' + str(row.row)] = row_values[10]
                ws['J' + str(row.row)] = row_values[11]
                ws['K' + str(row.row)] = row_values[12]
                ws['L' + str(row.row)] = row_values[13]
                ws['M' + str(row.row)] = row_values[14]
                row_values.pop(3)
                row_values.pop(3)

        elif row.value == "":

            for col in col_list:
                try:
                    cell = ws[col + str(row.row)].value.split(" ")
                    row_values.extend(cell)
                except AttributeError:
                    cell = ""
                    row_values.extend(cell)

            row_values = [x for x in row_values if x != ""]

            if "SUBTOTAL" in row_values:
                ws['N' + str(row.row)] = row_values[-1]
                for col in col_list:
                    ws[col + str(row.row)] = ""
                continue
            else:
                value_rows.append(row_values)

            if len(row_values) == 5 and row_values[0] != "TEXAS":
                for col in col_list:
                    ws[col + str(row.row)] = ""
                text = row_values[1] + " " + row_values[2]
                ws['F' + str(row.row)] = row_values[0]
                ws['I' + str(row.row)] = text
                ws['L' + str(row.row)] = row_values[3]
                ws['M' + str(row.row)] = row_values[4]

            if len(row_values) == 5 and row_values[0] == "TEXAS":
                for col in col_list:
                    ws[col + str(row.row)] = ""
                text = "TEXAS ENVIRONMENTAL TAX"
                ws['I' + str(row.row)] = text
                ws['L' + str(row.row)] = row_values[3]
                ws['M' + str(row.row)] = row_values[4]

            if len(row_values) == 4 and row_values[0] != "PRODUCTION":
                for col in col_list:
                    ws[col + str(row.row)] = ""
                ws['F' + str(row.row)] = row_values[0]
                ws['I' + str(row.row)] = row_values[1]
                ws['L' + str(row.row)] = row_values[2]
                ws['M' + str(row.row)] = row_values[3]

            if len(row_values) == 4 and row_values[0] == "PRODUCTION" or row_values[0] == "PIPELINE":
                for col in col_list:
                    ws[col + str(row.row)] = ""
                text = row_values[0] + " " + row_values[1]
                ws['I' + str(row.row)] = text
                ws['L' + str(row.row)] = row_values[2]
                ws['M' + str(row.row)] = row_values[3]

            if len(row_values) == 3:
                for col in col_list:
                    ws[col + str(row.row)] = ""
                ws['I' + str(row.row)] = row_values[0]
                ws['L' + str(row.row)] = row_values[1]
                ws['M' + str(row.row)] = row_values[2]

            if len(row_values) == 6 and row_values[1] == "TEXAS":
                for col in col_list:
                    ws[col + str(row.row)] = ""
                text = "TEXAS ENVIRONMENTAL TAX"
                ws['F' + str(row.row)] = row_values[0]
                ws['I' + str(row.row)] = text
                ws['L' + str(row.row)] = row_values[4]
                ws['M' + str(row.row)] = row_values[5]

            if "GRAND" in row_values:
                grand_total = row_values[-1].replace(",", "")
                for cell in row_values:
                    cell = None

    except TypeError:
        continue
ws.delete_rows(ws.max_row, 1)

# Turn date cells into date time
for cell in ws['A']:
    try:
        if len(cell.value) == 4:
            year = 2000 + int(cell.value[2:4])
            month = int(cell.value[0:2])
            day = 1
            date = datetime.date(year, month, day)
            cell.value = date
    except TypeError:
        continue

for cell in ws['B']:
    try:
        if cell.value == "ALL PLANT PRODUCTS":
            cell.value = "LIQUID PRODUCTS"
        elif "OIL" in cell.value:
            cell.value = "TEXAS OIL"
        elif "GAS" in cell.value:
            cell.value = "TEXAS GAS"
    except TypeError:
        continue

for cell in ws['C']:
    try:
        cell.value = int(cell.value)
    except(ValueError, TypeError):
        continue
# Turns all text based numbers into actual numbers
for col in full_col_list:
    active_column = ws[col]
    for cell in active_column:
        try:
            cell.value = cell.value.replace(',', "")
            cell.value = Decimal(cell.value)
        except(ValueError, TypeError, InvalidOperation, AttributeError):
            continue

# Finds all property tags
for cell in ws['A']:
    try:
        if "PROPERTY" in cell.value:
            tags.append(cell)
    except TypeError:
        continue

# Captures all cell values between property tags
for idx, cell in enumerate(tags):
    try:
        r = cell.row
        mr = (tags[idx + 1].row - 1)
        sheets.append(cell_range(r, mr))

    except IndexError:
        r = cell.row
        mr = ws.max_row
        sheets.append(cell_range(r, mr))

# Grab property name values
for tag in tags:
    sheet_names.append(ws['D' + str(tag.row)].value)

# Remove duplicate sheet names
new_sheet_list = []
[new_sheet_list.append(item) for item in sheet_names if item not in new_sheet_list]

# loop through sheets and combine all like properties
for sheet_name in new_sheet_list:
    combined_sheet = []
    for sheet in sheets:

        name = sheet[3].value

        if name == sheet_name:
            combined_sheet += sheet

    whole_sheets.append(combined_sheet)

sheet_names = [name.replace("NAME: ", "").replace(":", "")[0:30] for name in new_sheet_list]

# Create one new sheet per Property and fill in the values
for idx, sheet in enumerate(sheet_names):

    wb.create_sheet(sheet_names[idx])

    active_ws = wb[sheet_names[idx]]

    row_count = whole_sheets[idx][-1].row - whole_sheets[idx][0].row + 1
    ws['O' + str(whole_sheets[idx][-1].row)] = "=sum(N" + str(whole_sheets[idx][0].row + 1) + ":N" + str(whole_sheets[idx][-1].row)  + ")"

    i = 0
    for x in range(2, row_count + 2):
        for y in range(1, 14):
            active_ws.cell(row=x, column=y, value=whole_sheets[idx][i].value)
            i += 1

    # create column headings for each sheet
    for l, title in enumerate(headings):
        active_ws.cell(row=1, column=l + 1, value=title)

    # delete property names from sheets
    colA = active_ws["A"]
    for cell in colA:
        try:
            if "PROPERTY" in cell.value:
                active_ws.delete_rows(cell.row, 1)
        except TypeError:
            continue

for idx, sheet in enumerate(sheet_names):
    try:

        active_ws = wb[sheet_names[idx]]

        date_list = []
        items = []

        date_col = active_ws['A']

        for cell in date_col:
            if type(cell.value) is datetime.date:
                date_list.append(cell)

        for idx, cell in enumerate(date_list):
            sheet = []
            try:
                for cell_row in active_ws.iter_rows(min_row=cell.row, max_col=13, max_row=date_list[idx + 1].row - 1):
                    for data in cell_row:
                        sheet.append(data.value)

            except IndexError:
                for cell_row in active_ws.iter_rows(min_row=cell.row, max_col=13, max_row=active_ws.max_row):
                    for data in cell_row:
                        sheet.append(data.value)

            items.append(sheet)

        items = sorted(items, key=itemgetter(0, 1))

        flat_list = [item for sublist in items for item in sublist]

        i = 0
        for x in range(2, active_ws.max_row + 1):
            for y in range(1, 14):
                active_ws.cell(row=x, column=y, value=flat_list[i])
                i += 1
    except IndexError:
        continue

for idx, name in enumerate(sheet_names):
    active_ws = wb[sheet_names[idx]]
    active_ws['N' + str(active_ws.max_row)] = '=sum(M1:' + 'M' + str(active_ws.max_row + 2) + ')'
    sheet_total_index.append(active_ws['N' + str(active_ws.max_row)].coordinate)

wb.create_sheet('Totals', 0)
active_ws = wb['Totals']
# grab total values from sheets
for idx, name in enumerate(sheet_names):
    active_ws.cell(row=idx + 1, column=1, value=name)
    active_ws.cell(row=idx + 1, column=2, value="='" + sheet_names[idx] + "'!" + sheet_total_index[idx])
# checks total values
active_ws.cell(row=active_ws.max_row + 1, column=2, value="=sum(B1:B" + str(active_ws.max_row) + ")")
active_ws.cell(row=active_ws.max_row, column=3, value=Decimal(grand_total))
active_ws.cell(row=active_ws.max_row, column=4, value="=IF(C" + str(active_ws.max_row) + "=B" + str(
    active_ws.max_row) + ", \"True\", \"False\")")
active_ws['E' + str(active_ws.max_row)] = "=C" + str(active_ws.max_row) + "-" + "B" + str(active_ws.max_row)


# checks translated sheet totals against original sheet

for idx, sheet in enumerate(whole_sheets):
    active_ws['C' + str(idx + 1)] = "=sheet!O" + str(whole_sheets[idx][-1].row)
    active_ws['E' + str(idx + 1)] = "=C" + str(idx + 1) + "-" + "B" + str(idx + 1)



# save_file = input("What name would you like to save the file under?: ")
# print("Saving file...........................................")
# wb.save(save_file + '.xlsx')
# print('Save complete!')
#
#
# print("File saved to " + os.getcwd())

wb.save('swepi_clean.xlsx')

