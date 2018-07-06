import os
import openpyxl
from openpyxl import load_workbook
import csv
from openpyxl import Workbook
import datetime
import operator
from operator import itemgetter
from decimal import *

headings = ["REF", "CUSTOMER:JOB", "SHIP DATE", "DUE DATE", "ITEM", "GROSS VOL", "GROSS EXP",
            "GROSS OWNER"]

product_codes = {"100": " Texas Oil", "103": " Texas Oil", "300": "Condensate", "303": "Texas Oil", "200": "Texas Gas",
                 "201": "Texas Gas",
                 "202": "Texas Gas", "203": "Texas Gas", "204": "Texas Gas", "205": "Texas Gas", "206": "Texas Gas",
                 "209": "HELIUM", "400": "Liquid Products", "401": "Liquid Products", "402": "Liquid Products",
                 "403": "Liquid Products", "404": "Liquid Products", "405": "Liquid Products", "406": "Liquid Products",
                 "407": "Liquid Products", "40C": "Plant Condensate", "40D": "Plant Sulphur", "409": "Gasoline"}

col_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O']

needed_col_list = ['A', 'B', 'D', 'H', 'K', 'L', 'O']

row_list = []
tags = []
sheets = []
property_names = []
product_rows = []
only_value_rows = []
final_sort = []
interest_rows = []

os.chdir('/Users/monroe/Desktop')

now = datetime.datetime.now()

file_name = input("Please enter your file name: ")
print("Processing file...........................")


file = open(file_name, 'r')
csv_f = csv.reader(file)

with open("APCoutput.csv", "w", newline='') as f:
    writer = csv.writer(f, quoting=csv.QUOTE_NONNUMERIC)
    writer.writerows(csv_f)

wb = Workbook()
ws = wb.active
with open('APCoutput.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)

# Corrects Sheet formatting errors
for row in ws['A']:
    row_values = []
    if "PROPERTY" in row.value:

        for col in col_list:
            cell = ws[col + str(row.row)].value
            row_values.append(cell)

        row_values = list(filter(None, row_values))
        row_list.append(row_values)

        if len(row_values) == 4:
            ws['A' + str(row.row)] = row_values[0]
            ws['D' + str(row.row)] = row_values[1]
            ws['G' + str(row.row)] = row_values[2]
            ws['N' + str(row.row)] = row_values[3]
            ws['B' + str(row.row)] = ""
            ws['E' + str(row.row)] = ""
            ws['C' + str(row.row)] = ""
            ws['F' + str(row.row)] = ""
            ws['H' + str(row.row)] = ""
            ws['I' + str(row.row)] = ""
            ws['J' + str(row.row)] = ""
            ws['K' + str(row.row)] = ""

        elif len(row_values) == 2:
            ws['A' + str(row.row)] = row_values[0]
            ws['D' + str(row.row)] = row_values[1]
            ws['B' + str(row.row)] = ""
            ws['E' + str(row.row)] = ""
            ws['C' + str(row.row)] = ""
            ws['F' + str(row.row)] = ""
            ws['H' + str(row.row)] = ""
            ws['I' + str(row.row)] = ""
            ws['J' + str(row.row)] = ""
            ws['K' + str(row.row)] = ""

    else:
        for col in col_list:
            try:
                cell = ws[col + str(row.row)].value.split(" ")
                row_values.extend(cell)
            except AttributeError:
                cell = ""
                row_values.extend(cell)

        row_values = list(filter(None, row_values))
        row_list.append(row_values)

        if len(row_values) == 13:
            ws['A' + str(row.row)] = row_values[0]
            ws['B' + str(row.row)] = row_values[1]
            ws['C' + str(row.row)] = row_values[2]
            ws['D' + str(row.row)] = row_values[3]
            ws['E' + str(row.row)] = row_values[4]
            ws['F' + str(row.row)] = ""
            ws['G' + str(row.row)] = row_values[5]
            ws['H' + str(row.row)] = row_values[6]
            ws['I' + str(row.row)] = row_values[7]
            ws['J' + str(row.row)] = row_values[8]
            ws['K' + str(row.row)] = row_values[9]
            ws['L' + str(row.row)] = row_values[10]
            ws['M' + str(row.row)] = row_values[11]
            ws['N' + str(row.row)] = row_values[12]
        elif len(row_values) == 3:
            ws['H' + str(row.row)] = row_values[0]
            ws['L' + str(row.row)] = row_values[1]
            ws['M' + str(row.row)] = row_values[2]
            ws['E' + str(row.row)] = ""
            ws['I' + str(row.row)] = ""
            ws['J' + str(row.row)] = ""
            ws['G' + str(row.row)] = ""
            ws['F' + str(row.row)] = ""
            ws['K' + str(row.row)] = ""

        elif len(row_values) == 14:
            ws['A' + str(row.row)] = row_values[0]
            ws['B' + str(row.row)] = row_values[1]
            ws['C' + str(row.row)] = row_values[2]
            ws['D' + str(row.row)] = row_values[3]
            ws['E' + str(row.row)] = row_values[4]
            ws['F' + str(row.row)] = row_values[5]
            ws['G' + str(row.row)] = row_values[6]
            ws['H' + str(row.row)] = row_values[7]
            ws['I' + str(row.row)] = row_values[8]
            ws['J' + str(row.row)] = row_values[9]
            ws['K' + str(row.row)] = row_values[10]
            ws['L' + str(row.row)] = row_values[11]
            ws['M' + str(row.row)] = row_values[12]
            ws['N' + str(row.row)] = row_values[13]

        elif len(row_values) == 12:
            ws['A' + str(row.row)] = row_values[0]
            ws['B' + str(row.row)] = row_values[1]
            ws['C' + str(row.row)] = row_values[2]
            ws['D' + str(row.row)] = row_values[3]
            ws['E' + str(row.row)] = ""
            ws['F' + str(row.row)] = ""
            ws['G' + str(row.row)] = row_values[4]
            ws['H' + str(row.row)] = row_values[5]
            ws['I' + str(row.row)] = row_values[6]
            ws['J' + str(row.row)] = row_values[7]
            ws['K' + str(row.row)] = row_values[8]
            ws['L' + str(row.row)] = row_values[9]
            ws['M' + str(row.row)] = row_values[10]
            ws['N' + str(row.row)] = row_values[11]

        elif len(row_values) == 0:
            ws.delete_rows(row.row, 1)

        # else:
        #     print(row_values)

# Turns all text based numbers into actual numbers
for col in col_list:
    active_column = ws[col]
    for cell in active_column:
        try:
            if "(" in cell.value:
                cell.value = cell.value.replace("(", "").replace(")", "").replace(",", "")
                cell.value = Decimal(cell.value) * -1
            else:
                cell.value = cell.value.replace(',', "")
                cell.value = Decimal(cell.value)

        except(ValueError, TypeError, InvalidOperation, AttributeError):
            continue

# Turns dates into datetime objects
for cell in ws['A']:
    if "PROPERTY" not in cell.value and cell.value != "":
        try:
            year = 2000 + int(cell.value.split("/")[1])
            month = int(cell.value.split("/")[0])
            day = 1
            date = datetime.date(year, month, day)
            cell.value = date
        except IndexError:
            continue

# grab property names and codes
for cell in ws['A']:
    try:
        if "PROPERTY" in cell.value:
            tags.append(cell)
            property_names.append(ws['D' + str(cell.row)].value)
    except TypeError:
        continue

# Adds property names to end of rows
for idx, tag in enumerate(tags):
    try:
        row_count = tags[idx + 1].row - tag.row - 1
        for x in range(1, row_count + 1):
            ws['O' + str(tag.row + x)] = property_names[idx]
    except IndexError:
        row_count = ws.max_row - tag.row - 1
        for x in range(1, row_count + 1):
            ws['O' + str(tag.row + x)] = property_names[idx]

# Get row number for rows with products
for idx, tag in enumerate(tags):
    try:
        row_count = tags[idx + 1].row - tag.row - 1
        for x in range(1, row_count + 1):
            if ws['B' + str(tag.row + x)].value != "":
                product_rows.append(tag.row + x)
    except IndexError:
        row_count = ws.max_row - tag.row - 1
        for x in range(1, row_count + 1):
            product_rows.append(tag.row + x)

# eliminate empty rows
for idx, row in enumerate(product_rows):
    if ws['B' + str(row)].value == "":
        product_rows.pop(idx)

# read amended values back into lists
for row in range(1, ws.max_row + 1):
    cells = []
    for col in needed_col_list:
        cells.append(ws[col + str(row)].value)

    cells = [x for x in cells if x != ""]

    only_value_rows.append(cells)

# clean non-product rows out of list
only_value_rows = [item for item in only_value_rows if len(item) >= 6]


# special handling for interest rows
for row in ws['O']:
    try:
        if "INTEREST" in row.value:
            interest_row = list()
            interest_row.append(ws['A' + str(row.row)].value)
            interest_row.append(ws['B' + str(row.row)].value)
            interest_row.append(ws['N' + str(row.row)].value)
            interest_row.append(ws['O' + str(row.row)].value)
            if interest_row[0] != "":
                interest_rows.append(interest_row)
            else:
                continue
    except TypeError:
        continue

# make sure product code is str
for item in interest_rows:
    item[1] = str(item[1]).replace(".0", "")
# remove small unneeded rows
for idx, item in enumerate(interest_rows):
    if None in item:
        interest_rows.pop(idx)
# add to final sort list
final_sort.append(interest_rows)

# make product codes into ints if possible
for item in only_value_rows:
    try:
        item[1] = int(item[1])
    except ValueError:
        continue

# Sort value rows with the same date and same product code together
holder = list([only_value_rows[0]])
hold_counter = 0
for idx in range(1, len(only_value_rows) + 1):
    try:
        value_list = only_value_rows[idx]
    except IndexError:
        final_sort.append(holder)
        continue

    if value_list[0] == holder[hold_counter][0] and value_list[1] == holder[hold_counter][1] and holder[hold_counter][
        -1] == value_list[-1]:
        holder.append(value_list)
        hold_counter += 1
    else:
        final_sort.append(holder)
        holder = [value_list]
        hold_counter = 0

# ensure no empty lists in final sort
final_sort = [x for x in final_sort if x != []]



#
# total = Decimal(0)
# for cell in ws['N']:
#     if type(cell.value) == Decimal:
#         total += cell.value
#     print(total)


# Create new sheet in workbook
wb.create_sheet("totals")
active_ws = wb['totals']

# Create titles on sheet
for idx, title in enumerate(headings):
    active_ws.cell(row=1, column=idx + 1, value=title)

row_counter = 2

same_prop = ""

total = Decimal(0)
for item in final_sort:

    final_gross_vol = Decimal(0)
    final_exp = Decimal(0)
    final_owner_gross = Decimal(0)
    final_owner_adj = Decimal(0)

    if len(item) == 1 and "INTEREST" not in item:
        active_ws.cell(row=row_counter, column=2, value="APC:" + item[0][-1])  # Property Name
        active_ws.cell(row=row_counter + 1, column=2, value="APC:" + item[0][-1])  # Property Name
        active_ws.cell(row=row_counter, column=3, value=item[0][0])  # Date
        active_ws.cell(row=row_counter, column=4, value=item[0][0])  # Date
        active_ws.cell(row=row_counter, column=5, value=product_codes[str(item[0][1])])  # Product Code
        active_ws.cell(row=row_counter + 1, column=5, value="Texas Severence Exp")
        active_ws.cell(row=row_counter, column=6, value=item[0][2])  # Gross VOL
        active_ws.cell(row=row_counter + 1, column=7, value=item[0][3])  # Gross EXP
        active_ws.cell(row=row_counter, column=8, value=item[0][4])  # Gross Owner
        active_ws.cell(row=row_counter + 1, column=8, value=item[0][5] * -1)  # Owner Adj
        row_counter += 3

    else:
        for lst in item:
            if len(lst) > 4:
                gross_vol = Decimal(lst[2])
                final_gross_vol += Decimal(gross_vol)
                gross_exp = Decimal(lst[3])
                final_exp += Decimal(gross_exp)
                gross_owner = Decimal(lst[4])
                final_owner_gross += Decimal(gross_owner)
                owner_adj = Decimal(lst[5])
                final_owner_adj += Decimal(owner_adj)
            else:
                owner_gross = Decimal(lst[2])
                final_owner_gross += Decimal(owner_gross)

        active_ws.cell(row=row_counter, column=2, value="APC:" + item[0][-1])  # Property Name
        active_ws.cell(row=row_counter + 1, column=2, value="APC:" + item[0][-1])  # Property Name
        active_ws.cell(row=row_counter, column=3, value=item[0][0])  # Date
        active_ws.cell(row=row_counter, column=4, value=item[0][0])  # Date
        active_ws.cell(row=row_counter + 1, column=3, value=item[0][0])  # Date
        active_ws.cell(row=row_counter + 1, column=4, value=item[0][0])  # Date
        active_ws.cell(row=row_counter, column=5, value=product_codes[str(item[0][1])])  # Product Code
        active_ws.cell(row=row_counter + 1, column=5, value="Texas Severence Exp")
        active_ws.cell(row=row_counter, column=6, value=final_gross_vol)  # Gross VOL
        active_ws.cell(row=row_counter + 1, column=7, value=final_exp)  # Gross EXP
        active_ws.cell(row=row_counter, column=8, value=final_owner_gross)  # Gross Owner
        active_ws.cell(row=row_counter + 1, column=8, value=final_owner_adj * -1)  # Owner Adj
        row_counter += 3

save_file = input("What name would you like to save the file under?: ")

wb.save(save_file + '.xlsx')

print("File saved to " + os.getcwd())
